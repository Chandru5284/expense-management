from django.http import HttpResponse
from rest_framework.decorators import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
from datetime import datetime, timedelta
from openpyxl import Workbook
import tempfile
import os
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

# import models
from expense.models import Member, Transaction
from django.db.models import Sum

class DashboardTransactionView(APIView):
    def get(self, request, member_slug):
        try:
            # Get the current month's number and the previous month's number
            current_month = datetime.now().month
            previous_month = (datetime.now() - timedelta(days=30)).month
            
            # member model for filtering
            member = Member.objects.get(slug=member_slug)
            
            # Calculate total income for the current month
            current_month_income = Transaction.objects.filter(member=member, transaction_type__icontains="income", date__month=current_month).aggregate(total_income=Sum('amount'))
            total_income_this_month = current_month_income['total_income'] if current_month_income['total_income'] else 0
            
            # Calculate total income for the current month
            current_month_expense = Transaction.objects.filter(member=member, transaction_type__icontains="expense", date__month=current_month).aggregate(total_expense=Sum('amount'))
            total_expense_this_month = current_month_expense['total_expense'] if current_month_expense['total_expense'] else 0
            
            # total balance this month
            total_balance_this_month = total_income_this_month - total_expense_this_month
            
            # Calculate total income for the previous month
            previous_month_income = Transaction.objects.filter(member=member, transaction_type__icontains="income", date__month=previous_month).aggregate(total_income=Sum('amount'))
            total_income_last_month = previous_month_income['total_income'] if previous_month_income['total_income'] else 0
            
            # Calculate total income for the previous month
            previous_month_expense = Transaction.objects.filter(member=member, transaction_type__icontains="expense", date__month=previous_month).aggregate(total_expense=Sum('amount'))
            total_expense_last_month = previous_month_expense['total_expense'] if previous_month_expense['total_expense'] else 0
                        
            # total balance this month
            total_balance_last_month = total_income_last_month - total_expense_last_month
            
            # Calculate the income percentage change
            if total_income_last_month != 0:
                percentage_change = ((total_income_this_month - total_income_last_month) / total_income_last_month) * 100
            else:
                percentage_change = 100 
            
            income_percentage_change_formatted = "{:.2f}".format(percentage_change)
            
            # Calculate the expense percentage change
            if total_expense_last_month != 0:
                percentage_change = ((total_expense_this_month - total_expense_last_month) / total_expense_last_month) * 100
            else:
                percentage_change = 100 
            
            expense_percentage_change_formatted = "{:.2f}".format(percentage_change)
            
            # Calculate the expense percentage change
            if total_balance_last_month != 0:
                percentage_change = ((total_balance_this_month - total_balance_last_month) / total_balance_last_month) * 100
            else:
                percentage_change = 100 
            
            total_balance_percentage_change_formatted = "{:.2f}".format(percentage_change)
            
            this_month__total_transaction = Transaction.objects.filter(member=member, date__month=current_month).count()
            
            output = {
                "income": total_income_this_month,
                "expense": total_expense_this_month,
                "total_balance": total_balance_this_month,
                "income_percentage": income_percentage_change_formatted,
                "expense_percentage": expense_percentage_change_formatted,
                "total_balance_percentage": total_balance_percentage_change_formatted,
                "this_month__total_transaction": this_month__total_transaction
            }
            
            return Response(output, status=status.HTTP_200_OK)
        except Member.DoesNotExist:
            return Response({ "message" :  "Member record not exist"}, status=status.HTTP_404_NOT_FOUND)
        
        
class DashboardChartView(APIView):
    def get(self, request, member_slug):
        try:
            # Get the current year
            current_year = datetime.now().year
            
            # Define months
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            
            # member model for filtering
            member = Member.objects.get(slug=member_slug)
            
            # Initialize dictionary to store sales data
            income_data = {}
            expense_data = {}

            # Loop through each month to calculate sales
            for month in range(1, 13):
                # Calculate total sales for the month
                income = Transaction.objects.filter(member=member, transaction_type__icontains="income", date__year=current_year, date__month=month).aggregate(total_income=Sum('amount'))['total_income']
                expense = Transaction.objects.filter(member=member, transaction_type__icontains="expense", date__year=current_year, date__month=month).aggregate(total_expense=Sum('amount'))['total_expense']
                
                # Convert sales to "K" format for readability
                if income is not None:
                    income = round(income / 1000)  # Convert to thousands and round to 2 decimal places
                else:
                    income = 0
                    
                if expense is not None:
                    expense = round(expense / 1000)  # Convert to thousands and round to 2 decimal places
                else:
                    expense = 0
                    
                # Store sales data in dictionary
                income_data[months[month-1]] = income
                expense_data[months[month-1]] = expense
                
            output = {
                "income": income_data,
                "expense": expense_data
            }
            
            
            return Response(output, status=status.HTTP_200_OK)
        except Member.DoesNotExist:
            return Response({ "message" :  "Member record not exist"}, status=status.HTTP_404_NOT_FOUND)
        
        
class DownloadExcelView(APIView):
    permission_classes = (AllowAny, )
    def get(self, request, member_slug):
        try:
            
            query_params = {
                "start_date": request.GET.get('start_date', ''),
                "end_date": request.GET.get('end_date', ''),
                "transaction_type": request.GET.get('transaction_type', '')
            }
             
            # member model for filtering
            member = Member.objects.get(slug=member_slug)
            
            start_date_str = query_params['start_date']
            end_date_str = query_params['end_date']

            # Convert the start and end dates to datetime objects
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            # Query the sales data within the specified date range
            sales_data = Transaction.objects.filter(Q(member=member), Q(date__range=[start_date, end_date]), Q(transaction_type__icontains=query_params["transaction_type"])).order_by('date')
            
            print(query_params['start_date'], "-----", query_params['end_date'])

            # Create an Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Data"
            
            ws.merge_cells('A1:E1')
            ws.merge_cells('A2:E2')
        
            first_cell = ws['A1']
            first_cell.value = "Transaction List"
            first_cell.fill = PatternFill("solid", fgColor="246ba1")
            first_cell.font  = Font(bold=True, color="F7F6FA")
            first_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            second_cell = ws['A2']
            second_cell.value = "("+"From: " + start_date_str + " To: " + end_date_str+")"
            second_cell.fill = PatternFill("solid", fgColor="EF4444")
            second_cell.font  = Font(bold=True, color="F7F6FA")
            second_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            
            titles = ["Date", "Category", "Transaction type", "Title", "Amount"]
            row_num = 3
            
            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(titles, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.fill = PatternFill("solid", fgColor="50C878")
                cell.font  = Font(bold=True, color="F7F6FA")           
            
            for sale in sales_data:
                print(sales_data)
                ws.append([sale.date.strftime('%Y-%m-%d'), sale.category.title, sale.transaction_type, sale.title, sale.amount])

            # Apply borders
            border_style = Side(style='thin', color='000000')  # You can customize style and color
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            # Iterate through each cell and apply the border
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
        
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_index = column[0].column
                column = [cell.value for cell in column]
                for cell in column:
                    try:
                        if len(str(cell)) > max_length:
                            max_length = len(cell)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[get_column_letter(column_index)].width = adjusted_width
                        
            # Create a temporary file to save the workbook
            with tempfile.NamedTemporaryFile(delete=False) as tmp:
                file_path = tmp.name
                wb.save(file_path)

            # Open the file for reading in binary mode
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'

            # Delete the temporary file
            os.unlink(file_path)      
            
            return response
        except Member.DoesNotExist:
            return Response({ "message" :  "Member record not exist"}, status=status.HTTP_404_NOT_FOUND)